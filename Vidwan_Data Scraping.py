from selenium import webdriver

from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
path = input("Enter the path of your excel file without " " (Where list of id is written) :")
# path ="E:\Python Programming\Vidwan Portal Datahandling project\Vidwan ID Created till 04_03_2023.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

list_Name=[]
list_Designation=[]
list_Department=[]
list_Organisation=[]
list_Vidwan_Score = []
list_Articles=[]
list_Awards=[]
list_Books=[]
list_Projects=[]
list_Publications=[]
list_Co_author=[]
list_Conference_Proceedings=[]
list_Citations=[]
list_H_index=[]
list_Crossref_citations=[]
list_Altmetrices=[]
list_Facebook=[]
list_Twitter=[]
list_Crossref=[]
list_Google_Scholar=[]
list_Google_scholar_citations=[]
list_Google_Scholar_H_index=[]
list_Google_Scholar_i10_index=[]
list_Orcid_id=[]
list_Scopus_id=[]
list_Researcher_id=[]
list_Google_Scholar_id=[]

list_Orcid_EMP=[]
list_Orcid_work=[]
list_Scopus_citations=[]
list_Scopus_doc=[]
list_Scopus_hindex=[]
list_Expertise=[]
list_WebOfScience_hindex=[]
list_PublicationInWebOfScience=[]
list_SumOfTimeOfCited_wos=[]
list_Citing_Articles_wos=[]
for i in range(2,sheet.max_row+1):
    
    cellbox = sheet.cell(row = i, column = 9)
    v_id = cellbox.value
    
    # driver = webdriver.Chrome("D:\chromedriver-win64 (1)\chromedriver-win64\chromedriver.exe")
    driver_path = input("Enter the path of your ChromeDriver: ")
    driver = webdriver.Chrome(driver_path)
    driver.get("https://vidwan.inflibnet.ac.in/profile/"+str(v_id))
    content = driver.page_source
    soup = BeautifulSoup(content)
    
    data_dict = {"v_score":" ",
                  "Articles":" ",
                  "Books":" ",
                  "Awards":" ",
                  "Name":" ",
                  "Designation":" ",
                  "Organisation":" ",
                  "Conf_proceeding":" ",
                  "Citation":" ",
                  "h_index":" ",
                  "crossref_citation":" ",
                  "Altmetrices":" ",
                  "Facebook":" ",
                  "Twitter":" ",
                  "Crossref":" ",
                  "Google_scholar":" ",
                  "g_citation":" ",
                  "g_h_index":" ",
                  "g_i10_index":" ",
                  "publications":" ",
                  "co_author":" ",
                  "Department":" ",
                  "Orcid_id":" ",
                  "Scopus_id":" ",
                  "Researcher_id":" ",
                  "Google_scholar_id":" ",
                  "Projects":" ",
                  "Orcid_EMP":" ",
                  "Orcid_work":" ",
                  "Scopus_citations":" ",
                  "Scopus_doc":" ",
                  "Scopus_hindex":" ",
                  "Expertise":" ",
                  "WebOfScience_hindex":" ",
                  "PublicationInWebOfScience":" ",
                  "SumOfTimeOfCited_wos":" ",
                  "Citing_Articles_wos":" "}
    if v_id =="":
        pass
    else:
        for a in soup.find_all("div",class_="statistics"):
            v_score = a.find("span",class_="pull-right")
            data_dict["v_score"] = v_score.text
            
        for b in soup.find_all("div",class_="service-block-v3 service-block-blue"):
            list1 = b.find_all("span",class_="counter")
            try:
                data_dict["Articles"] = list1[0].text
            except:
                pass
            try:
                data_dict["Books"] = list1[1].text
            except:
                pass
            try:
                data_dict["Projects"] = list1[2].text
            except:
                pass
            try:
                data_dict["Awards"] = list1[3].text
            except:
                pass
            
            
            #hwe jyare blue box update thy, or koy iitan na prof ma projects ni details add thy tyare programm ni valuse missmatch thy jy, to dynaic banava mate nu logic develop kr.
            '''
            list2 = b.find_all("span",class_="service-heading")
            for i in list1:
                for j in list2:
                    if i == j:
            '''
            
        for c in soup.find_all("div",class_="profile-bio"):
            list2 = c.find_all("span",class_="col-sm-12")
            data_dict["Name"] = list2[0].text
            data_dict["Designation"] = list2[1].text  
            data_dict["Organisation"] = list2[2].text
           
        for d in soup.find_all("div",class_="profile_articles_part"):
            list3 = d.find_all("li",class_="Pub_li_br_dashed")
            try:
                conf_p = list3[2].find("div",class_="counter p0")
                data_dict["Conf_proceeding"] = conf_p.text
            except:
                pass
        
        temp_loop_var = 0
        for e in soup.find_all("div",class_="service-block-v3 p0"):
            for e1 in e.find_all("div",class_="col-md-4 col-sm-4 col-xs-12 mb10-xs"):
                    if temp_loop_var == 0:
                        citation = e1.find("span",class_="counter")
                        if len(citation.text) > 0:
                            data_dict["Citation"] = citation.text
                        else:
                            pass
                        temp_loop_var = 1
                    crossref_citation = e1.find("span",class_="counter")
                    crf_c = re.findall('[0-9]+',crossref_citation.text)
                   # print(crf_c)
                    if len(crf_c) > 0:   
                        data_dict["crossref_citation"] = crf_c[0]
                    else:
                       pass
                
        for f in soup.find_all("div",class_="col-md-4 col-sm-4 col-xs-12 mb10-xs"):
            for f1 in f.find_all("div",class_="Cell-citation br1"):
                h_index = f1.find("span",class_="counter")
                data_dict["h_index"] = h_index.text
                
        #for single Almetrics column....     
        '''       
        for g in soup.find_all("div",class_="panel body"):
            for g1 in g.find_all("div",class_="Cell-citation"):
                Altmetrices = g1.find("span",class_="counter")
                data_dict["Altmetrices"] = Altmetrices.text
        '''
        try:
            for g in soup.find_all("div",class_="service-block-v3 p0"):
                list4 = g.find_all("div",class_="Cell-citation br1")
                try:
                    Altmetrices = list4[0].find("span",class_="counter")
                    data_dict["Altmetrices"] = Altmetrices.text
                except:
                    pass
            try:
                Facebook = list4[1].find("span",class_="counter")  
                data_dict["Facebook"] = Facebook.text 
            except:
                pass
            try:
                Twitter = list4[2].find("span",class_="counter")
                data_dict["Twitter"] = Twitter.text
            except:
                pass
            try:
                Crossref = list4[3].find("span",class_="counter")
                data_dict["Crossref"] = Crossref.text
            except:
                pass
            try:
                Google_scholar = list4[4].find("span",class_="counter")
                data_dict["Google_scholar"] = Google_scholar.text
            except:
                pass
        except:
            pass
        
        for h in soup.find_all("h2",class_="panel-title heading-sm pull-right"):
            list5 = h.find_all("button",class_="btn-u badge-blue")
            pub = re.findall('[0-9]+',list5[0].text)
            co_a = re.findall('[0-9]+',list5[1].text)
            data_dict["publications"] = pub[0]
            data_dict["co_author"] = co_a[0]
        
        for i in soup.find_all("div",class_="tag-box tag-box-v1 margin-bottom-40"):
            try:
                dept = i.find("span",id="e_expertise")
                data_dict["Department"] = dept.text
            except:
                pass
            
        for j in soup.find_all("div",class_="col-md-3 md-margin-bottom-40"):
            for j1 in j.find_all("div",id="identity-view"):
                for j2 in j1.find_all("div",class_="mCSB_container mCS_y_hidden mCS_no_scrollbar_y"):
                    try:
                        Orcid_id = j2.find("span",id="i_orcid_id")
                        data_dict["Orcid_id"] = Orcid_id.text
                    except:
                        pass
                
                    try:
                        Scopus_id = j2.find("span",id="i_scopus_id")
                        data_dict["Scopus_id"] = Scopus_id.text
                    except:
                        pass
                    
                    try:
                        Researcher_id = j2.find("span",id="i_isi_id")
                        data_dict["Researcher_id"] = Researcher_id.text
                    except:
                        pass
                    
                    try:
                        Google_scholar_id = j2.find("span",id="i_google_sid")
                        data_dict["Google_scholar_id"] = Google_scholar_id.text
                    except:
                        pass
        
            
        for j in soup.find_all("div",class_="col-md-3 md-margin-bottom-40"):
            for j1 in j.find_all("div",id="identity-view"):
                for j2 in j1.find_all("div",class_="mCSB_container mCS_y_hidden mCS_no_scrollbar_y"):
                    try:
                        Orcid_id = j2.find("span",id="i_orcid_id")
                        data_dict["Orcid_id"] = Orcid_id.text
                        try:
                            driver.get("https://orcid.org/"+Orcid_id.text)
                            orcid_content = driver.page_source
                            o_soup = BeautifulSoup(orcid_content)
                            try:
                                for i in o_soup.find_all("section",id="affiliations"):
                                    for j in i.find_all("h3",class_="activity-header orc-font-body-large"):
                                        emp = j.find("span",class_="clickable ng-star-inserted")
                                        emp = re.findall('[0-9]+',emp.text)
                                        data_dict["Orcid_EMP"] = emp[0]
                            except:
                                pass
                            
                            try:
                                for i1 in o_soup.find_all("section",id="works"):
                                    for jj in i1.find_all("h3",class_="activity-header orc-font-body-large"):
                                        work = jj.find("span",class_="clickable ng-star-inserted")
                                        work = re.findall('[0-9]+',work.text)
                                        data_dict["Orcid_work"] = work[0]
                            except:
                                pass
                        except:
                            pass
                    except:
                        pass
                
                    try:
                        Scopus_id = j2.find("span",id="i_scopus_id")
                        data_dict["Scopus_id"] = Scopus_id.text
                        driver.get("https://www.scopus.com/authid/detail.uri?authorId="+Scopus_id.text)
                        try:
                            scopus_content = driver.page_source
                            s_soup = BeautifulSoup(scopus_content)
                            try:
                                scopus_data = []
                                for i in s_soup.find_all("section",class_="MetricSection-module__3xldN"):
                                    for j in i.find_all("div",class_="verticalHighlightColor_96e11d"):
                                        sc = j.find("span",class_="typography_f0ad1e font-size-xl_f0ad1e sans_f0ad1e")
                                        sc = re.findall('[0-9]+',sc.text)
                                        scopus_data.append(sc[0])
                                data_dict["Scopus_citations"] = scopus_data[0]
                                data_dict["Scopus_doc"] = scopus_data[1]
                                data_dict["Scopus_hindex"] = scopus_data[2]
                            except:
                                pass
                        except:
                            pass
                    except:
                        pass
                    
                    try:
                        Researcher_id = j2.find("span",id="i_isi_id")
                        data_dict["Researcher_id"] = Researcher_id.text
                        driver.get("https://www.webofscience.com/wos/author/record/"+Researcher_id.text)
                        time.sleep(5)
                        try:
                            button = driver.find_element(By.ID, 'dismissGuides')
                            button.click()
                            wos_content = driver.page_source
                            wos_soup = BeautifulSoup(wos_content)
                            try:
                                for m in wos_soup.find_all("div",class_="wat-author-metric-main-div ng-star-inserted"):
                                        lst = m.find_all("div",class_="wat-author-metric")
                                        data_dict["WebOfScience_hindex"]=lst[0].text
                                        data_dict["PublicationInWebOfScience"]=lst[1].text
                                        data_dict["SumOfTimeOfCited_wos"]=lst[2].text
                                        try:
                                            for j in m.find_all("div",class_="wat-author-metric-inline-block wat-author-metric-left-padding wat-author-record__metrics-item-column"):
                                                wos_citing_articles = j.find("span",class_="wat-author-metric ng-star-inserted")
                                                data_dict["Citing_Articles_wos"]=wos_citing_articles.text
                                        except:
                                            pass
                            except:
                                pass
                        except:
                            pass
                    except:
                        pass

                    try:
                        Google_scholar_id = j2.find("span",id="i_google_sid")
                        data_dict["Google_scholar_id"] = Google_scholar_id.text
                        driver.get("https://scholar.google.co.in/citations?user="+Google_scholar_id.text)
                        try:
                            g_content = driver.page_source
                            g_soup = BeautifulSoup(g_content)
                            try:
                                google_data=[]
                                for i in g_soup.find_all("div",id="gsc_rsb_cit"):
                                    tbl = i.find("table")
                                    for j in tbl.find_all('td'):
                                       gd = j.text
                                       google_data.append(gd)
                                data_dict["g_citation"]=google_data[1]
                                data_dict["g_h_index"]=google_data[4]
                                data_dict["g_i10_index"]=google_data[7]
                            except:
                                pass
                            try:
                                exp=[]
                                expertise_string=""
                                for k in g_soup.find_all("div",id="gsc_prf_i"):
                                    list_expertise = k.find_all("a",class_="gsc_prf_inta gs_ibl")
                                    for k2 in range(0,len(list_expertise)):
                                        exp.append(list_expertise[k2].text)
                                    for k3 in range(0,len(exp)):
                                        expertise_string = expertise_string + ", " + exp[k3] 
                                    expertise_string = expertise_string[2:]
                                data_dict["Expertise"]=expertise_string
                            except:
                                pass
                        except:
                            pass
                    except:
                        pass
        
        list_Name.append(data_dict["Name"])
        list_Designation.append(data_dict["Designation"])
        list_Department.append(data_dict["Department"])
        list_Organisation.append(data_dict["Organisation"])
        list_Vidwan_Score.append(data_dict["v_score"])
        list_Articles.append(data_dict["Articles"])
        list_Awards.append(data_dict["Awards"])
        list_Books.append(data_dict["Books"])
        list_Projects.append(data_dict["Projects"])
        list_Publications.append(data_dict["publications"])
        list_Co_author.append(data_dict["co_author"])
        list_Conference_Proceedings.append(data_dict["Conf_proceeding"])
        list_Citations.append(data_dict["Citation"])
        list_H_index.append(data_dict["h_index"])
        list_Crossref_citations.append(data_dict["crossref_citation"])
        list_Altmetrices.append(data_dict["Altmetrices"])
        list_Facebook.append(data_dict["Facebook"])
        list_Twitter.append(data_dict["Twitter"])
        list_Crossref.append(data_dict["Crossref"])
        list_Google_Scholar.append(data_dict["Google_scholar"])
        list_Google_scholar_citations.append(data_dict["g_citation"])
        list_Google_Scholar_H_index.append(data_dict["g_h_index"])
        list_Google_Scholar_i10_index.append(data_dict["g_i10_index"])
        list_Orcid_id.append(data_dict["Orcid_id"])
        list_Scopus_id.append(data_dict["Scopus_id"])
        list_Researcher_id.append(data_dict["Researcher_id"])
        list_Google_Scholar_id.append(data_dict["Google_scholar_id"])
        
        list_Orcid_EMP.append(data_dict["Orcid_EMP"])
        list_Orcid_work.append(data_dict["Orcid_work"])
        list_Scopus_citations.append(data_dict["Scopus_citations"])
        list_Scopus_doc.append(data_dict["Scopus_doc"])
        list_Scopus_hindex.append(data_dict["Scopus_hindex"])
        list_Expertise.append(data_dict["Expertise"])
        list_WebOfScience_hindex.append(data_dict["WebOfScience_hindex"])
        list_PublicationInWebOfScience.append(data_dict["PublicationInWebOfScience"])
        list_SumOfTimeOfCited_wos.append(data_dict["SumOfTimeOfCited_wos"])
        list_Citing_Articles_wos.append(data_dict["Citing_Articles_wos"])
        

df_dict = {"Name":" ","Designation":" ","Department":" ","Organisation":" ","Vidwan Score":" ",
             "Articles":" ","Awards":" ","Books":" ","Projects":" ","Conference Proceedings":" ",
             "Citations":" ","H-index":" ","Crossref citations":" ","Altmetrics_Newspaper":" ","Altmetrics_Facebook":" ",
             "Altmetrics_Twitter":" ","Altmetrics_mendeley":" ","Altmetrics_GooglePlus":" ","Google scholar citations":" ",
             "Google scholar H-index":" ","Google scholar i-10 index":" ","Orcid ID":" ","Scopus ID":" ",
             "Researcher ID":" ", "Google Scholar ID":" ","Publications":" ","NoofCoAuthors":" ",
             "Employeement (Orcid)":" ","No. of works in orcid":" ","Scopus Citations":" ","Scopus documents":" ",
             "Scopus H-index":" ","Expertises (As per the google scholar)":" ","Web of Science H-index":" ",
             "Web of Science Publications":" ","Sum of time cited (Web of science)":" ","Citing articles (Web of science)":" "}

df_dict["Name"]=list_Name
df_dict["Designation"]=list_Designation
df_dict["Department"]=list_Department
df_dict["Organisation"]=list_Organisation
df_dict["Vidwan Score"]=list_Vidwan_Score
df_dict["Articles"]=list_Articles
df_dict["Awards"]=list_Awards
df_dict["Books"]=list_Books
df_dict["Projects"]=list_Projects
df_dict["Conference Proceedings"]=list_Conference_Proceedings
df_dict["Citations"]=list_Citations
df_dict["H-index"]=list_H_index
df_dict["Crossref citations"]=list_Crossref_citations
df_dict["Altmetrics_Newspaper"]=list_Altmetrices
df_dict["Altmetrics_Facebook"]=list_Facebook
df_dict["Altmetrics_Twitter"]=list_Twitter
df_dict["Altmetrics_mendeley"]=list_Crossref
df_dict["Altmetrics_GooglePlus"]=list_Google_Scholar
df_dict["Google scholar citations"]=list_Google_scholar_citations
df_dict["Google scholar H-index"]=list_Google_Scholar_H_index
df_dict["Google scholar i-10 index"]=list_Google_Scholar_i10_index
df_dict["Orcid ID"]=list_Orcid_id
df_dict["Scopus ID"]=list_Scopus_id
df_dict["Researcher ID"]=list_Researcher_id
df_dict["Google Scholar ID"]=list_Google_Scholar_id
df_dict["Publications"]=list_Publications
df_dict["NoofCoAuthors"]=list_Co_author

df_dict["Employeement (Orcid)"]=list_Orcid_EMP
df_dict["No. of works in orcid"]=list_Orcid_work
df_dict["Scopus Citations"]=list_Scopus_citations
df_dict["Scopus documents"]=list_Scopus_doc
df_dict["Scopus H-index"]=list_Scopus_hindex
df_dict["Expertises (As per the google scholar)"]=list_Expertise
df_dict["Web of Science H-index"]=list_WebOfScience_hindex
df_dict["Web of Science Publications"]=list_PublicationInWebOfScience
df_dict["Sum of time cited (Web of science)"]=list_SumOfTimeOfCited_wos
df_dict["Citing articles (Web of science)"]=list_Citing_Articles_wos


df = pd.DataFrame(data=df_dict)
df.to_excel("Vidwan_Profile_Details_WebScraped_through_Python.xlsx", index=False)

